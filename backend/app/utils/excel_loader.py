import os
from pathlib import Path
from openpyxl import load_workbook
from app.models.database import get_db

BASE_DIR = Path(__file__).resolve().parents[2]
RAW_DATA_DIR = BASE_DIR / "data" / "raw"

PRODUCTS_FILE = RAW_DATA_DIR / "products.xlsx"
ORDERS_FILE = RAW_DATA_DIR / "orders.xlsx"


def load_products():
    if not PRODUCTS_FILE.exists():
        print("Products file not found.")
        return

    conn = get_db()
    cursor = conn.cursor()

    # Skip if already loaded
    cursor.execute("SELECT COUNT(*) as count FROM medicines")
    if cursor.fetchone()["count"] > 0:
        conn.close()
        print("Medicines already loaded. Skipping.")
        return

    wb = load_workbook(PRODUCTS_FILE)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row or all(cell is None for cell in row):
            continue

        try:
            product_id, name, pzn, price, package_size, description = row
        except ValueError:
            continue

        if not product_id or not name:
            continue

        try:
            price = float(price) if price is not None else 0.0
        except (ValueError, TypeError):
            price = 0.0

        # -------------------------
        # Prescription Logic (Demo Rule-Based)
        # -------------------------
        prescription_required = "No"

        if any(keyword in name for keyword in [
            "Ramipril",
            "Minoxidil",
            "femiLoges"
        ]):
            prescription_required = "Yes"

        cursor.execute("""
            INSERT OR IGNORE INTO medicines 
            (product_id, name, pzn, price, package_size, description, stock, prescription_required)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            product_id,
            name,
            str(pzn) if pzn else None,
            price,
            package_size,
            description,
            100,  # default demo stock
            prescription_required
        ))

    conn.commit()
    conn.close()
    print("Products imported successfully.")


def load_orders():
    if not ORDERS_FILE.exists():
        print("Orders file not found.")
        return

    conn = get_db()
    cursor = conn.cursor()

    # Skip if already loaded
    cursor.execute("SELECT COUNT(*) as count FROM orders")
    if cursor.fetchone()["count"] > 0:
        conn.close()
        print("Orders already loaded. Skipping.")
        return

    wb = load_workbook(ORDERS_FILE)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row or all(cell is None for cell in row):
            continue

        try:
            (
                patient_id,
                age,
                gender,
                purchase_date,
                product_name,
                quantity,
                total_price,
                dosage_frequency,
                prescription_required
            ) = row
        except ValueError:
            continue

        if not patient_id or not product_name or quantity is None or total_price is None:
            continue

        try:
            quantity = int(quantity)
            total_price = float(total_price)
        except (ValueError, TypeError):
            continue

        # Insert customer
        cursor.execute("""
            INSERT OR IGNORE INTO customers (id, age, gender)
            VALUES (?, ?, ?)
        """, (patient_id, age, gender))

        # Insert order
        cursor.execute("""
            INSERT INTO orders (
                customer_id,
                product_name,
                quantity,
                purchase_date,
                total_price,
                dosage_frequency,
                prescription_required
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            patient_id,
            product_name,
            quantity,
            str(purchase_date),
            total_price,
            dosage_frequency,
            prescription_required
        ))

    conn.commit()
    conn.close()
    print("Orders imported successfully.")


def load_all_data():
    load_products()
    load_orders()